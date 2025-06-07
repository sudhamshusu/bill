import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

def create_excel_template(data_sheet, boq_sheet):
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])
    
    # Add Data sheet
    ws_data = wb.create_sheet("Data")
    for row in data_sheet.iterrows():
        for col_idx, value in enumerate(row[1], 1):
            ws_data.cell(row=row[0]+1, column=col_idx, value=value)
    
    # Add BOQ sheet
    ws_boq = wb.create_sheet("BOQ")
    for row in boq_sheet.iterrows():
        for col_idx, value in enumerate(row[1], 1):
            ws_boq.cell(row=row[0]+1, column=col_idx, value=value)
    
    # Apply formatting to Data sheet
    format_data_sheet(ws_data)
    
    # Apply formatting to BOQ sheet
    format_boq_sheet(ws_boq)
    
    # Create Cover Page sheet
    create_cover_page(wb, ws_data)
    
    # Create IPC Summary sheet
    create_ipc_summary(wb, ws_data, ws_boq)
    
    # Create measurement sheets (1-1, 1-2, etc.)
    create_measurement_sheets(wb, ws_data, ws_boq)
    
    return wb

def format_data_sheet(ws):
    # Apply formatting to Data sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['F'].width = 30

def format_boq_sheet(ws):
    # Apply formatting to BOQ sheet
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Set column widths
    column_widths = {
        'A': 5, 'B': 60, 'C': 8, 'D': 10, 'E': 10, 'F': 12,
        'G': 10, 'H': 10, 'I': 12, 'J': 10, 'K': 12, 'L': 10,
        'M': 12, 'N': 10, 'O': 12, 'P': 10, 'Q': 12, 'R': 10, 'S': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Apply borders to data area
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=18, max_row=43, min_col=1, max_col=19):
        for cell in row:
            cell.border = thin_border

def create_cover_page(wb, ws_data):
    ws_cover = wb.create_sheet("cover page")
    
    # Add government structure
    govt_structure = [
        "Government",
        "Ministry",
        "Department",
        "FRSMO",
        "Office",
        "District"
    ]
    
    for i, item in enumerate(govt_structure, start=4):
        ws_cover.cell(row=i, column=1, value=f"={item}")
    
    # Add project info from Data sheet
    ws_cover['A28'] = f"=Data!B5"  # Name of Project
    ws_cover['A29'] = f"=Data!B8"  # Contract Identification No
    ws_cover['A42'] = f"=Data!B4"  # Bill No.
    
    # Add submitted by/to sections
    ws_cover['A54'] = "Submitted By:"
    ws_cover['C54'] = "Submitted to:"
    ws_cover['A55'] = f"=Data!B10"  # Name of Contractor
    ws_cover['C55'] = f"=Data!F9"   # Client
    
    # Formatting
    ws_cover.column_dimensions['A'].width = 30
    ws_cover.column_dimensions['C'].width = 30
    
    for row in ws_cover.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')

def create_ipc_summary(wb, ws_data, ws_boq):
    ws_ipc = wb.create_sheet("IPC Summary")
    
    # Add headers
    ws_ipc['A8'] = "Contract Bill Summary"
    ws_ipc['A9'] = f"=Data!B5"  # Name of Project
    ws_ipc['A10'] = f"=Data!B10"  # Name of Contractor
    ws_ipc['A11'] = f"=Data!B8"  # Contract Identification No
    ws_ipc['A12'] = f"=Data!B7"  # Project Implement place
    ws_ipc['A13'] = "IPC No.: First"
    
    # Add table headers
    headers = [
        "S.N.", "Works Description", 
        "Total Amount Provision in Original Contract (NRs.)",
        "Total Amount Provision in Revised Contract (NRs.)",
        "Upto Date Bill Amount (NRs.)", "Previous Bill Amount (NRs.)",
        "Present Bill Amount (NRs.)", "Remaining Amount",
        "Progress W.R.T Revised Amount"
    ]
    
    for col, header in enumerate(headers, start=1):
        ws_ipc.cell(row=15, column=col, value=header)
    
    # Add data rows with formulas
    rows = [
        ("A", "General works", 
         f"=SUM(BOQ!F19:F20)", f"=SUM(BOQ!K19:K20)", 
         f"=SUM(BOQ!K19:K20)", f"=BOQ!M23", 
         f"=SUM(BOQ!O19:O20)", f"=C16-E16", f"=E16/C16"),
        ("B", f"=BOQ!B22", f"=BOQ!F23", "", f"=BOQ!K23", 
         "0", f"=SUM(BOQ!O23)", f"=C17-E17", f"=E17/C17"),
        ("C", f"=BOQ!B25", f"=SUM(BOQ!F26:F37)", "", 
         f"=SUM(BOQ!K26:K37)", "0", f"=SUM(BOQ!O26:O37)", 
         f"=C18-E18", f"=E18/C18"),
        ("D", f"=BOQ!B35", f"=SUM(BOQ!F36:F37)", 
         f"=SUM(BOQ!O23:O37)", f"=SUM(BOQ!K36:K37)", 
         "0", f"=SUM(BOQ!O36:O37)", f"=C19-E19", f"=E19/C19")
    ]
    
    for row_idx, row_data in enumerate(rows, start=16):
        for col_idx, value in enumerate(row_data, start=1):
            ws_ipc.cell(row=row_idx, column=col_idx, value=value)
    
    # Add totals
    ws_ipc['A20'] = "Total (A+B+C+D)"
    ws_ipc['C20'] = f"=SUM(C16:C19)"
    ws_ipc['D20'] = f"=SUM(D16:D19)"
    ws_ipc['E20'] = f"=SUM(E16:E19)"
    ws_ipc['F20'] = f"=SUM(F16:F19)"
    ws_ipc['G20'] = f"=SUM(G16:G19)"
    ws_ipc['H20'] = f"=C20-E20"
    ws_ipc['I20'] = f"=E20/C20"
    
    # Add VAT and Grand Total
    ws_ipc['A21'] = "VAT@13%"
    ws_ipc['C21'] = f"=TRUNC((C20-BOQ!F19)*0.13,2)"
    ws_ipc['D21'] = f"=TRUNC((D20-BOQ!I19)*0.13,2)"
    ws_ipc['E21'] = f"=TRUNC((E20-BOQ!K19)*0.13,2)"
    ws_ipc['F21'] = f"=TRUNC(F20*0.13,2)"
    ws_ipc['G21'] = f"=TRUNC((G20-BOQ!O19)*0.13,2)"
    ws_ipc['H21'] = f"=C21-E21"
    ws_ipc['I21'] = f"=E21/C21"
    
    ws_ipc['A22'] = "Grand Total"
    ws_ipc['C22'] = f"=C20+C21"
    ws_ipc['D22'] = f"=D20+D21"
    ws_ipc['E22'] = f"=E20+E21"
    ws_ipc['F22'] = f"=F20+F21"
    ws_ipc['G22'] = f"=G20+G21"
    ws_ipc['H22'] = f"=C22-E22"
    ws_ipc['I22'] = f"=E22/C22"
    
    # Formatting
    ws_ipc.column_dimensions['A'].width = 5
    ws_ipc.column_dimensions['B'].width = 40
    ws_ipc.column_dimensions['C'].width = 15
    ws_ipc.column_dimensions['D'].width = 15
    ws_ipc.column_dimensions['E'].width = 15
    ws_ipc.column_dimensions['F'].width = 15
    ws_ipc.column_dimensions['G'].width = 15
    ws_ipc.column_dimensions['H'].width = 15
    ws_ipc.column_dimensions['I'].width = 15
    
    for row in ws_ipc.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Apply borders to data area
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws_ipc.iter_rows(min_row=15, max_row=22, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border

def create_measurement_sheets(wb, ws_data, ws_boq):
    # Create measurement sheets for each BOQ item
    measurement_items = [
        ("1-1", 19, "PS"), ("1-2", 20, "m3"), ("2-1", 23, "m3"),
        ("3-1", 26, "m3"), ("3-2", 27, "m3"), ("3-3", 28, "m3"),
        ("3-4", 29, "m3"), ("3-5", 30, "m3"), ("3-6", 31, "m3"),
        ("3-7", 32, "m3"), ("3-8", 33, "m3"), ("4-1", 36, "m3"),
        ("4-2", 37, "m3")
    ]
    
    for sheet_name, boq_row, unit in measurement_items:
        ws = wb.create_sheet(sheet_name)
        
        # Add header rows (similar to 1-1 sheet)
        for i in range(1, 8):
            ws.cell(row=i, column=1, value=f"='1-1'!A{i}:I{i}")
        
        ws['A8'] = "Measurement Sheet"
        
        for i in range(9, 11):
            ws.cell(row=i, column=1, value=f"='1-1'!A{i}:I{i}")
        
        # Add table headers
        if sheet_name == "4-2":
            # Special format for 4-2 (reinforcement)
            headers = [
                "S.N.", "Description of works", "", "Main bar", "", "", "", 
                "Distrinution Bar", "", "", "", "Dowel Bar", "", "", "", 
                "Tie Bar", "", "", "", "Total Quantity Q = Q1+Q2+Q3+Q4", "Remarks"
            ]
            subheaders = [
                "", "", "Unit", "No.", "Length (m)", "Dia (m)", "Unit wt (kg)", 
                "Quantity (Q1)", "No.", "Length (m)", "Dia (m)", "Unit wt (kg)", 
                "Quantity (Q2)", "No.", "Length (m)", "Dia (m)", "Unit wt (kg)", 
                "Quantity (Q3)", "No.", "Length (m)", "Dia (m)", "Unit wt (kg)", 
                "Quantity (Q4)", ""
            ]
            
            for col, header in enumerate(headers, start=1):
                ws.cell(row=13, column=col, value=header)
            
            for col, header in enumerate(subheaders, start=1):
                ws.cell(row=14, column=col, value=header)
            
            # Add data row
            ws['A15'] = f"=BOQ!A{boq_row}"
            ws['B15'] = f"=BOQ!B{boq_row}"
            ws['C15'] = "m3"
            
            # Add summary rows
            ws['T16'] = "Total Quantity in KG"
            ws['X16'] = f"=SUM(X15:X15)"
            
            ws['T17'] = "Total Quantity MT"
            ws['X17'] = f"=TRUNC(X16/1000,2)"
            
            ws['T18'] = "Pervious Bill Quantity"
            ws['X18'] = f"=BOQ!L{boq_row}"
            
            ws['T19'] = "This Bill Quantity"
            ws['X19'] = f"=X17"
            
            # Merge cells for better formatting
            ws.merge_cells('A13:A14')
            ws.merge_cells('B13:B14')
            ws.merge_cells('C13:C14')
            ws.merge_cells('V13:W13')
            ws.merge_cells('V14:W14')
            
        else:
            # Standard format for other sheets
            if sheet_name == "3-2":
                headers = [
                    "S.N.", "Description of works", "Unit", "No.", "Length (m)", 
                    "Breadth b1 (m)", "Breadth b2 (m)", "Average Breadth B=(b1+b2)/2 (m)", 
                    "Height (m)", "Quantity", "Remarks", "IPCs"
                ]
            elif sheet_name == "3-4" or sheet_name == "3-5":
                headers = [
                    "S.N.", "Description of works", "Unit", "No.", "Length (L)", 
                    "Base", "Top", "Avg Breadth (B)", "Height (H)", 
                    "Quantity Q = L*B*H", "Remarks", "IPCs"
                ]
            elif sheet_name == "4-1":
                headers = [
                    "", "Description of works", "Unit", "", "Length (m)", 
                    "Breadth (b1)", "Breadth (b2)", "Breadth B = (b1+b2)/2", 
                    "Thickness (H)", "Quantity (Q ) = L*B*H", "Remarks", "IPCs"
                ]
            else:
                headers = [
                    "S.N.", "Description of works", "Unit", "No.", "Length (m)", 
                    "Breadth (m)", "Height (m)", "Quantity", "Remarks"
                ]
            
            for col, header in enumerate(headers, start=1):
                ws.cell(row=12, column=col, value=header)
            
            # Add data row
            ws['A13'] = f"=BOQ!A{boq_row}"
            ws['B13'] = f"=BOQ!B{boq_row}"
            ws['C13'] = unit
            
            # Add summary rows
            if sheet_name in ["1-1", "1-2", "3-6", "3-7", "3-8"]:
                ws['E19'] = "Total Quantity"
                ws['H19'] = f"=H12"
                
                ws['E20'] = "Pervious Bill Quantity"
                ws['H20'] = f"=BOQ!L{boq_row}"
                
                ws['E21'] = "This Bill Quantity"
                ws['H21'] = f"=H19-H20"
            elif sheet_name in ["2-1"]:
                ws['E17'] = "Total Quantity"
                ws['H17'] = f"=SUM(H14:H15)"
                
                ws['E18'] = "Previous bill Quantity"
                ws['H18'] = f"=BOQ!L25"
                
                ws['E19'] = "This Bill Quantity"
                ws['H19'] = f"=H17-H18"
            elif sheet_name in ["3-1"]:
                ws['E34'] = "Total Quantity"
                ws['H34'] = f"=SUM(H15:H33)"
                
                ws['E35'] = "Pervious Bill Quantity"
                ws['H35'] = f"=BOQ!L26"
                
                ws['E36'] = "This Bill Quantity"
                ws['H36'] = f"=H34-H35"
            elif sheet_name in ["3-2"]:
                ws['E29'] = "Total Quantity"
                ws['J29'] = f"=SUM(J15:J28)"
                
                ws['E30'] = "Pervious Bill Quantity"
                ws['J30'] = f"=BOQ!L27"
                
                ws['E31'] = "This Bill Quantity"
                ws['J31'] = f"=J29-J30"
            elif sheet_name in ["3-3"]:
                ws['E25'] = "Total Quantity"
                ws['H25'] = f"=SUM(H15:H24)"
                
                ws['E26'] = "Pervious Bill Quantity"
                ws['H26'] = f"=BOQ!L28"
                
                ws['E27'] = "This Bill Quantity"
                ws['H27'] = f"=H25-H26"
            elif sheet_name in ["3-4", "3-5"]:
                ws['E22'] = "Total Quantity"
                ws['J22'] = f"=SUM(J16:J21)"
                
                ws['E23'] = "Pervious Bill Quantity"
                ws['J23'] = f"=BOQ!L{boq_row}"
                
                ws['E24'] = "This Bill Quantity"
                ws['J24'] = f"=J22-J23"
            elif sheet_name in ["4-1"]:
                ws['G28'] = "Total Quantity"
                ws['J28'] = f"=SUM(J15:J27)"
                
                ws['G29'] = "Pervious bill Quantity"
                ws['J29'] = f"=BOQ!L36"
                
                ws['G30'] = "This Bill Quantity"
                ws['J30'] = f"=J28-J29"
        
        # Formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Set column widths
        if sheet_name == "4-2":
            for col in range(1, 25):
                ws.column_dimensions[get_column_letter(col)].width = 8
        else:
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 5
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            if sheet_name in ["3-2", "3-4", "3-5", "4-1"]:
                ws.column_dimensions['I'].width = 10
                ws.column_dimensions['J'].width = 10
                ws.column_dimensions['K'].width = 10
                ws.column_dimensions['L'].width = 10

def main():
    st.title("Bill Generation App")
    st.write("Upload your Data and BOQ sheets to generate the complete bill structure")
    
    # File uploaders
    data_file = st.file_uploader("Upload Data Sheet (Excel)", type=["xlsx"])
    boq_file = st.file_uploader("Upload BOQ Sheet (Excel)", type=["xlsx"])
    
    if data_file and boq_file:
        try:
            # Read the uploaded files
            data_sheet = pd.read_excel(data_file, sheet_name="Data", header=None)
            boq_sheet = pd.read_excel(boq_file, sheet_name="BOQ", header=None)
            
            # Generate the complete workbook
            with st.spinner("Generating bill structure..."):
                wb = create_excel_template(data_sheet, boq_sheet)
                
                # Save to BytesIO buffer
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Download button
                st.success("Bill generated successfully!")
                st.download_button(
                    label="Download Generated Bill",
                    data=output,
                    file_name="generated_bill.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
